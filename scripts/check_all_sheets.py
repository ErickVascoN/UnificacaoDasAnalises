"""
Verifica TODAS as abas da planilha de Lençol para encontrar OPs 81 e 82
"""

import requests
import io
import pandas as pd
from openpyxl import load_workbook

LENCOL_SPREADSHEET_ID = "1ypSEpTvIsm_hbgHmEf-v0fuR-P9h0mOa"

url_xlsx = f"https://docs.google.com/spreadsheets/d/{LENCOL_SPREADSHEET_ID}/export?format=xlsx"

print("=" * 80)
print("DIAGNÓSTICO: Verificando todas as abas do XLSX")
print("=" * 80)

try:
    print(f"\n📥 Baixando XLSX...")
    r = requests.get(url_xlsx, timeout=60)
    r.raise_for_status()
    
    xlsx_file = io.BytesIO(r.content)
    
    # Carrega workbook
    wb = load_workbook(xlsx_file)
    print(f"\n📋 Abas encontradas: {wb.sheetnames}")
    
    # Verifica cada aba
    for sheet_name in wb.sheetnames:
        print(f"\n{'=' * 80}")
        print(f"ABA: {sheet_name}")
        print(f"{'=' * 80}")
        
        try:
            df = pd.read_excel(xlsx_file, sheet_name=sheet_name, header=0)
            print(f"  Dimensões: {len(df)} linhas × {len(df.columns)} colunas")
            print(f"  Colunas: {list(df.columns)}")
            
            # Procura OPs 81 e 82
            for op_num in ["81", "82"]:
                matches = df[df.astype(str).apply(lambda x: x.str.contains(op_num, na=False)).any(axis=1)]
                if not matches.empty:
                    print(f"\n  ✓ OP {op_num} encontrada ({len(matches)} linha(s)):")
                    for idx, row in matches.iterrows():
                        print(f"    {row.to_dict()}")
                else:
                    print(f"\n  ✗ OP {op_num} não encontrada")
            
            # Mostra todas as OPs com data 27/05
            if "DATA" in df.columns:
                df_27_05 = df[df["DATA"].astype(str).str.contains("27/05|5/27", na=False)]
                if not df_27_05.empty:
                    print(f"\n  📅 Linhas com 27/05:")
                    for idx, row in df_27_05.iterrows():
                        if "OP" in df.columns:
                            print(f"    OP: {row.get('OP', '?')} | {row.to_dict()}")
                        else:
                            print(f"    {row.to_dict()}")
        
        except Exception as e:
            print(f"  ❌ Erro ao ler aba: {e}")
    
except Exception as e:
    print(f"❌ Erro ao baixar XLSX: {e}")

print("\n" + "=" * 80)
