r"""
Gera o arquivo Excel  data/Producao_Semanal_Externas.xlsx
com 4 abas: MEGA BARIRI | PREVITTEX MATRIZ | PREVITTEX FILIAL | MEGA PREVEN

Depois e so abrir o Google Drive, fazer upload e "Abrir com Google Planilhas".

Execute:
    venv\Scripts\python.exe scripts\criar_template_externas.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).parent.parent / "data" / "Producao_Semanal_Externas.xlsx"

# ── Dados por fábrica: (PRODUTO, CLIENTE) ────────────────────────────────────────
FABRICAS: dict[str, dict] = {
    "MEGA BARIRI": {
        "cor_header": "0D4F2E",   # verde escuro
        "exemplos": [
            ("COBERTOR TOQUE DE SEDA", "NC INDUSTRIA"),
            ("MANTA",                  "NC INDUSTRIA"),
        ],
    },
    "PREVITTEX MATRIZ": {
        "cor_header": "1A3373",   # azul escuro
        "exemplos": [
            ("MANTA BABY",      "ANDREZA"),
            ("JOGOS DE CAMA",   "ANDREZA"),
            ("FRONHAS",         "CAMESA"),
            ("VELOUR",          "CAMESA"),
            ("MANTA PRENSADA",  "CAMESA"),
            ("MANTA C/ CINTA",  "CAMESA"),
            ("COBERTOR 180G",   "CAMESA"),
            ("BABY",            "CAMESA"),
            ("CORTINA",         "CAMESA"),
            ("JOGO DE CAMA",    "CORTTEX"),
        ],
    },
    "PREVITTEX FILIAL": {
        "cor_header": "1E4080",   # azul médio
        "exemplos": [
            ("MANTA BABY",    "ANDREZA"),
            ("JOGOS DE CAMA", "ANDREZA"),
            ("CORTINA",       "DECOR"),
            ("JOGO DE CAMA",  "DECOR"),
        ],
    },
    "MEGA PREVEN": {
        "cor_header": "5C1A5C",   # roxo escuro
        "exemplos": [
            ("FRONHAS",               "MARCELINO"),
            ("JG DUPLO PONTO PALITO", "MARCELINO"),
            ("FRONHA PONTO PALITO",   "MARCELINO"),
            ("TONHAS",                "MARCELINO"),
        ],
    },
}

HEADERS = ["DATA", "PRODUTO", "CLIENTE", "QUANTIDADE"]
COL_WIDTHS = [14, 30, 20, 14]   # caracteres por coluna

# ── Estilos ──────────────────────────────────────────────────────────────────────
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FONT_HEADER  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_EXAMPLE = Font(name="Calibri", italic=True, color="888888", size=10)
FONT_DATA    = Font(name="Calibri", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _build_sheet(wb: Workbook, nome: str, config: dict) -> None:
    ws = wb.create_sheet(title=nome)
    cor  = config["cor_header"]
    rows = config["exemplos"]

    fill_header  = _fill(cor)
    fill_example = _fill("F2F2F2")

    # ── Cabeçalho ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = FONT_HEADER
        cell.fill      = fill_header
        cell.border    = BORDER
        cell.alignment = ALIGN_CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"   # congela cabeçalho

    # ── Linhas de exemplo (referência — a fábrica substitui pelos dados reais) ──
    for row_idx, (produto, cliente) in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 18
        values = ["DD/MM/AAAA", produto, cliente, 0]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = FONT_EXAMPLE
            cell.fill      = fill_example
            cell.border    = BORDER
            cell.alignment = ALIGN_CENTER if col_idx in (1, 4) else ALIGN_LEFT

    # ── Linhas em branco para preenchimento (50 linhas extras) ───────────────────
    next_row = len(rows) + 2
    for row_idx in range(next_row, next_row + 50):
        ws.row_dimensions[row_idx].height = 18
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.font      = FONT_DATA
            cell.border    = BORDER
            cell.alignment = ALIGN_CENTER if col_idx in (1, 4) else ALIGN_LEFT


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)   # remove a aba vazia padrão

    for nome, config in FABRICAS.items():
        _build_sheet(wb, nome, config)
        print(f"  [OK] Aba criada: {nome}")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nArquivo salvo em:\n   {OUTPUT}")
    print("\nPROXIMOS PASSOS:")
    print("  1. Abra o Google Drive")
    print("  2. Faca upload deste arquivo")
    print("  3. Clique com botao direito -> 'Abrir com' -> 'Google Planilhas'")
    print("  4. Copie o ID da URL e informe ao desenvolvedor para conectar ao dashboard")
    print("  5. Compartilhe cada aba com o e-mail da fabrica correspondente (permissao Editor)\n")


if __name__ == "__main__":
    main()
